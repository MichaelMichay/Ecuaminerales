from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import models
from .models import Insumo, CategoriaInsumo, LugarConsumo, MovimientoInventario
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from .models import Insumo, CategoriaInsumo, LugarConsumo
from .forms import InsumoForm, CategoriaInsumoForm, LugarConsumoForm
from usuarios.decorators import rol_requerido
from usuarios.models import Usuario
from reportes.models import Notificacion, Auditoria
from django.core.paginator import Paginator
from django.contrib import messages
import csv
from reportlab.lib import colors
from django.http import HttpResponse
from django.conf import settings
import os
from datetime import datetime

@login_required
@rol_requerido(['Administrador', 'Bodeguero'])
def lista_insumos(request):
    buscar = request.GET.get('buscar')
    estado_stock = request.GET.get('estado_stock')

    insumos = Insumo.objects.all().order_by('nombre_insumo')

    if buscar:
        insumos = insumos.filter(nombre_insumo__icontains=buscar)

    if estado_stock == 'bajo':
        insumos = insumos.filter(stock__lte=models.F('stock_minimo'))

    if estado_stock == 'normal':
        insumos = insumos.filter(stock__gt=models.F('stock_minimo'))

    paginator = Paginator(insumos, 10)
    page_number = request.GET.get('page')
    insumos = paginator.get_page(page_number)

    return render(request, 'inventario/lista_insumos.html', {
        'insumos': insumos,
        'buscar': buscar,
        'estado_stock': estado_stock,
        'page_obj': insumos
    })

@login_required
@rol_requerido(['Administrador', 'Bodeguero'])
def crear_insumo(request):
    if request.method == 'POST':
        form = InsumoForm(request.POST)

        if form.is_valid():
            insumo = form.save()

            Auditoria.objects.create(
                usuario=request.user,
                accion='CREACIÓN DE INSUMO',
                descripcion=f'Se creó el insumo {insumo.nombre_insumo}'
            )

            if insumo.stock <= insumo.stock_minimo:
                usuarios_notificar = Usuario.objects.filter(
                    rol__nombre_rol__in=['Administrador', 'Bodeguero']
                )

                for usuario in usuarios_notificar:
                    Notificacion.objects.create(
                        usuario=usuario,
                        mensaje=f'El insumo {insumo.nombre_insumo} está en stock bajo.'
                    )

            messages.success(request, 'Insumo creado correctamente.')

            return redirect('inventario')

        else:
            messages.error(request, 'No se pudo crear el insumo. Revisa los datos.')

    else:
        form = InsumoForm()

    return render(request, 'inventario/crear_insumo.html', {
        'form': form
    })

@login_required
@rol_requerido(['Administrador', 'Bodeguero'])
def editar_insumo(request, id):

    insumo = Insumo.objects.get(id=id)

    stock_anterior = insumo.stock

    if request.method == 'POST':

        form = InsumoForm(request.POST, instance=insumo)

        if form.is_valid():

            insumo_editado = form.save(commit=False)

            stock_nuevo = insumo_editado.stock

            diferencia = stock_nuevo - stock_anterior

            insumo_editado.save()

            if diferencia > 0:

                MovimientoInventario.objects.create(
                    tipo_movimiento='ENTRADA',
                    cantidad=diferencia,
                    stock_anterior=stock_anterior,
                    stock_actual=stock_nuevo,
                    observacion='Ajuste manual de inventario: aumento de stock.',
                    insumo=insumo_editado,
                    usuario=request.user
                )

            elif diferencia < 0:

                MovimientoInventario.objects.create(
                    tipo_movimiento='SALIDA',
                    cantidad=abs(diferencia),
                    stock_anterior=stock_anterior,
                    stock_actual=stock_nuevo,
                    observacion='Ajuste manual de inventario: disminución de stock.',
                    insumo=insumo_editado,
                    usuario=request.user
                )

            messages.success(
                request,
                'Insumo actualizado correctamente y movimiento registrado en Kardex.'
            )

            return redirect('inventario')

        else:

            messages.error(
                request,
                'No se pudo actualizar el insumo. Revisa los datos.'
            )

    else:

        form = InsumoForm(instance=insumo)

    return render(request, 'inventario/editar_insumo.html', {
        'form': form,
        'insumo': insumo
    })


@login_required
@rol_requerido(['Administrador', 'Bodeguero'])
def eliminar_insumo(request, id):
    insumo = Insumo.objects.get(id=id)

    Auditoria.objects.create(
        usuario=request.user,
        accion='ELIMINACIÓN DE INSUMO',
        descripcion=f'Se eliminó el insumo {insumo.nombre_insumo}'
    )

    insumo.delete()

    messages.success(request, 'Insumo eliminado correctamente.')

    return redirect('inventario')

@login_required
@rol_requerido(['Administrador', 'Bodeguero'])
def reporte_inventario_pdf(request):

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    insumos = Insumo.objects.all().order_by('nombre_insumo')

    def encabezado():

        logo_path = os.path.join(
            settings.BASE_DIR,
            'static',
            'img',
            'logo_ecuaminerales.jpeg'
        )

        if os.path.exists(logo_path):
            p.drawImage(
                logo_path,
                40,
                height - 90,
                width=75,
                height=55,
                preserveAspectRatio=True,
                mask='auto'
            )

        p.setFillColor(colors.HexColor('#243b80'))
        p.setFont('Helvetica-Bold', 22)

        p.drawCentredString(
            width / 2,
            height - 45,
            'ECUMINERALES S.A.'
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica-Bold', 15)

        p.drawCentredString(
            width / 2,
            height - 70,
            'Reporte de Inventario de Explosivos'
        )

        p.setFillColor(colors.HexColor('#555555'))
        p.setFont('Helvetica-Oblique', 9)

        p.drawCentredString(
            width / 2,
            height - 88,
            'Sistema de Control de Explosivos e Inventario'
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 9)

        p.drawString(
            40,
            height - 115,
            f'Generado por: {request.user.username}'
        )

        p.drawRightString(
            width - 40,
            height - 115,
            f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        )

        p.setStrokeColor(colors.HexColor('#243b80'))
        p.setLineWidth(1)
        p.line(40, height - 130, width - 40, height - 130)

    def cabecera_tabla(y):

        p.setFillColor(colors.HexColor('#243b80'))
        p.rect(40, y - 6, width - 80, 22, fill=True, stroke=False)

        p.setFillColor(colors.white)
        p.setFont('Helvetica-Bold', 8)

        p.drawString(45, y, 'NOMBRE')
        p.drawString(165, y, 'TIPO')
        p.drawString(235, y, 'STOCK')
        #p.drawString(285, y, 'MIN.')
        p.drawString(335, y, 'UNIDAD')
        #p.drawString(405, y, 'PELIGRO')
        p.drawString(485, y, 'ESTADO')

        p.setFillColor(colors.black)

    def pie_pagina():

        p.setFont('Helvetica', 8)
        p.setFillColor(colors.grey)

        p.drawString(
            40,
            40,
            'ECUMINERALES S.A. - Sistema de control de explosivos e inventario'
        )

        p.drawRightString(
            width - 40,
            40,
            f'Página {p.getPageNumber()}'
        )

    encabezado()

    y = height - 165

    cabecera_tabla(y)

    y -= 28

    p.setFont('Helvetica', 8)

    contador = 0

    for insumo in insumos:

        estado = 'STOCK BAJO' if insumo.stock <= insumo.stock_minimo else 'NORMAL'

        if contador % 2 == 0:
            p.setFillColor(colors.HexColor('#f2f4f8'))
            p.rect(40, y - 5, width - 80, 20, fill=True, stroke=False)

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 8)

        p.drawString(45, y, str(insumo.nombre_insumo)[:23])
        p.drawString(165, y, str(insumo.get_tipo_insumo_display())[:12])
        p.drawRightString(265, y, str(insumo.stock))
        #p.drawRightString(315, y, str(insumo.stock_minimo))
        p.drawString(335, y, str(insumo.unidad_medida)[:12])
        #p.drawString(405, y, str(insumo.nivel_peligrosidad)[:12])

        if estado == 'STOCK BAJO':
            p.setFillColor(colors.HexColor('#dc3545'))
        else:
            p.setFillColor(colors.HexColor('#198754'))

        p.roundRect(485, y - 4, 65, 14, 4, fill=True, stroke=False)

        p.setFillColor(colors.white)
        p.setFont('Helvetica-Bold', 7)

        p.drawCentredString(
            517,
            y,
            estado
        )

        p.setFont('Helvetica', 8)
        p.setFillColor(colors.black)

        y -= 22
        contador += 1

        if y < 80:

            pie_pagina()

            p.showPage()

            encabezado()

            y = height - 165

            cabecera_tabla(y)

            y -= 28

            p.setFont('Helvetica', 8)

    pie_pagina()

    p.showPage()
    p.save()

    return response

@login_required
@rol_requerido(['Administrador'])
def lista_categorias(request):
    categorias = CategoriaInsumo.objects.all().order_by('nombre_categoria')

    paginator = Paginator(categorias, 10)
    page_number = request.GET.get('page')
    categorias = paginator.get_page(page_number)

    return render(request, 'inventario/lista_categorias.html', {
        'categorias': categorias,
        'page_obj': categorias
    })

@login_required
@rol_requerido(['Administrador'])
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaInsumoForm(request.POST)

        if form.is_valid():
            categoria = form.save()

            Auditoria.objects.create(
                usuario=request.user,
                accion='CREACIÓN DE CATEGORÍA',
                descripcion=f'Se creó la categoría {categoria.nombre_categoria}'
            )

            messages.success(request, 'Categoría creada correctamente.')
            return redirect('categorias')

        else:
            messages.error(request, 'No se pudo crear la categoría. Revisa los datos.')

    else:
        form = CategoriaInsumoForm()

    return render(request, 'inventario/crear_categoria.html', {
        'form': form
    })
@login_required
@rol_requerido(['Administrador'])
def editar_categoria(request, id):
    categoria = CategoriaInsumo.objects.get(id=id)

    if request.method == 'POST':
        form = CategoriaInsumoForm(request.POST, instance=categoria)

        if form.is_valid():
            categoria = form.save()

            Auditoria.objects.create(
                usuario=request.user,
                accion='EDICIÓN DE CATEGORÍA',
                descripcion=f'Se editó la categoría {categoria.nombre_categoria}'
            )

            messages.success(request, 'Categoría actualizada correctamente.')
            return redirect('categorias')

        else:
            messages.error(request, 'No se pudo actualizar la categoría. Revisa los datos.')

    else:
        form = CategoriaInsumoForm(instance=categoria)

    return render(request, 'inventario/editar_categoria.html', {
        'form': form,
        'categoria': categoria
    })

@login_required
@rol_requerido(['Administrador'])
def eliminar_categoria(request, id):
    categoria = CategoriaInsumo.objects.get(id=id)

    Auditoria.objects.create(
        usuario=request.user,
        accion='ELIMINACIÓN DE CATEGORÍA',
        descripcion=f'Se eliminó la categoría {categoria.nombre_categoria}'
    )

    categoria.delete()

    messages.success(request, 'Categoría eliminada correctamente.')
    return redirect('categorias')
@rol_requerido(['Administrador'])
def lista_lugares(request):
    lugares = LugarConsumo.objects.all().order_by('nombre')

    paginator = Paginator(lugares, 10)
    page_number = request.GET.get('page')
    lugares = paginator.get_page(page_number)

    return render(request, 'inventario/lista_lugares.html', {
        'lugares': lugares,
        'page_obj': lugares
    })

@login_required
@rol_requerido(['Administrador'])
def crear_lugar(request):
    if request.method == 'POST':
        form = LugarConsumoForm(request.POST)

        if form.is_valid():
            lugar = form.save()

            Auditoria.objects.create(
                usuario=request.user,
                accion='CREACIÓN DE LUGAR',
                descripcion=f'Se creó el lugar de consumo {lugar.nombre}'
            )

            messages.success(request, 'Lugar de consumo creado correctamente.')
            return redirect('lugares')

        else:
            messages.error(request, 'No se pudo crear el lugar. Revisa los datos.')

    else:
        form = LugarConsumoForm()

    return render(request, 'inventario/crear_lugar.html', {
        'form': form
    })

@login_required
@rol_requerido(['Administrador'])
def editar_lugar(request, id):
    lugar = LugarConsumo.objects.get(id=id)

    if request.method == 'POST':
        form = LugarConsumoForm(request.POST, instance=lugar)

        if form.is_valid():
            lugar = form.save()

            Auditoria.objects.create(
                usuario=request.user,
                accion='EDICIÓN DE LUGAR',
                descripcion=f'Se editó el lugar de consumo {lugar.nombre}'
            )

            messages.success(request, 'Lugar de consumo actualizado correctamente.')
            return redirect('lugares')

        else:
            messages.error(request, 'No se pudo actualizar el lugar. Revisa los datos.')

    else:
        form = LugarConsumoForm(instance=lugar)

    return render(request, 'inventario/editar_lugar.html', {
        'form': form,
        'lugar': lugar
    })
@login_required
@rol_requerido(['Administrador'])
def eliminar_lugar(request, id):
    lugar = LugarConsumo.objects.get(id=id)

    Auditoria.objects.create(
        usuario=request.user,
        accion='ELIMINACIÓN DE LUGAR',
        descripcion=f'Se eliminó el lugar de consumo {lugar.nombre}'
    )

    lugar.delete()

    messages.success(request, 'Lugar de consumo eliminado correctamente.')
    return redirect('lugares')

@login_required
@rol_requerido(['Administrador', 'Bodeguero'])
def kardex_inventario(request):

    buscar = request.GET.get('buscar')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    movimientos = MovimientoInventario.objects.all().order_by('-fecha_movimiento')

    if buscar:
        movimientos = movimientos.filter(
            insumo__nombre_insumo__icontains=buscar
        )

    if fecha_inicio:
        movimientos = movimientos.filter(
            fecha_movimiento__date__gte=fecha_inicio
        )

    if fecha_fin:
        movimientos = movimientos.filter(
            fecha_movimiento__date__lte=fecha_fin
        )

    paginator = Paginator(movimientos, 10)
    page_number = request.GET.get('page')
    movimientos = paginator.get_page(page_number)

    return render(request, 'inventario/kardex.html', {
        'movimientos': movimientos,
        'buscar': buscar,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'page_obj': movimientos
    })

@login_required
@rol_requerido(['Administrador', 'Bodeguero'])
def kardex_pdf(request):

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="kardex_inventario.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    movimientos = MovimientoInventario.objects.all().order_by('-fecha_movimiento')

    def encabezado():

        logo_path = os.path.join(
            settings.BASE_DIR,
            'static',
            'img',
            'logo_ecuaminerales.jpeg'
        )

        if os.path.exists(logo_path):
            p.drawImage(
                logo_path,
                40,
                height - 90,
                width=75,
                height=55,
                preserveAspectRatio=True,
                mask='auto'
            )

        p.setFillColor(colors.HexColor('#243b80'))
        p.setFont('Helvetica-Bold', 22)

        p.drawCentredString(
            width / 2,
            height - 45,
            'ECUAMINERALES S.A.'
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica-Bold', 15)

        p.drawCentredString(
            width / 2,
            height - 70,
            'Kardex de Inventario'
        )

        p.setFillColor(colors.HexColor('#555555'))
        p.setFont('Helvetica-Oblique', 9)

        p.drawCentredString(
            width / 2,
            height - 88,
            'Historial de entradas y salidas de insumos'
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 9)

        p.drawString(
            40,
            height - 115,
            f'Generado por: {request.user.username}'
        )

        p.drawRightString(
            width - 40,
            height - 115,
            f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        )

        p.setStrokeColor(colors.HexColor('#243b80'))
        p.setLineWidth(1)
        p.line(40, height - 130, width - 40, height - 130)

    def cabecera_tabla(y):

        p.setFillColor(colors.HexColor('#243b80'))
        p.rect(30, y - 6, width - 60, 22, fill=True, stroke=False)

        p.setFillColor(colors.white)
        p.setFont('Helvetica-Bold', 7)

        p.drawString(35, y, 'FECHA')
        p.drawString(110, y, 'USUARIO')
        p.drawString(175, y, 'INSUMO')
        p.drawString(285, y, 'TIPO')
        p.drawString(350, y, 'CANT.')
        p.drawString(405, y, 'ANTERIOR')
        p.drawString(470, y, 'ACTUAL')
        p.drawString(525, y, 'OBS.')

        p.setFillColor(colors.black)

    def pie_pagina():

        p.setFont('Helvetica', 8)
        p.setFillColor(colors.grey)

        p.drawString(
            40,
            40,
            'ECUAMINERALES S.A. - Sistema de control de explosivos e inventario'
        )

        p.drawRightString(
            width - 40,
            40,
            f'Página {p.getPageNumber()}'
        )

    encabezado()

    y = height - 165
    cabecera_tabla(y)
    y -= 28

    contador = 0

    for mov in movimientos:

        if contador % 2 == 0:
            p.setFillColor(colors.HexColor('#f2f4f8'))
            p.rect(30, y - 5, width - 60, 22, fill=True, stroke=False)

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 7)

        p.drawString(
            35,
            y,
            mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M')
        )

        usuario = mov.usuario.username if mov.usuario else 'Sin usuario'
        p.drawString(110, y, usuario[:12])

        insumo = mov.insumo.nombre_insumo if mov.insumo else 'Sin insumo'
        p.drawString(175, y, insumo[:18])

        tipo = str(mov.tipo_movimiento)

        if tipo == 'ENTRADA':
            color_tipo = '#198754'
        else:
            color_tipo = '#dc3545'

        p.setFillColor(colors.HexColor(color_tipo))
        p.roundRect(285, y - 4, 55, 14, 4, fill=True, stroke=False)

        p.setFillColor(colors.white)
        p.setFont('Helvetica-Bold', 6)
        p.drawCentredString(312, y, tipo)

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 7)

        p.drawRightString(380, y, str(mov.cantidad))
        p.drawRightString(450, y, str(mov.stock_anterior))
        p.drawRightString(510, y, str(mov.stock_actual))

        observacion = mov.observacion if mov.observacion else ''
        p.drawString(525, y, str(observacion)[:12])

        y -= 24
        contador += 1

        if y < 80:

            pie_pagina()
            p.showPage()

            encabezado()

            y = height - 165
            cabecera_tabla(y)
            y -= 28

    pie_pagina()

    p.showPage()
    p.save()

    return response

@login_required
@rol_requerido(['Administrador', 'Bodeguero'])
def kardex_excel(request):

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kardex Inventario'

    azul = '243B80'
    verde = '198754'
    rojo = 'DC3545'
    gris = 'F2F4F8'
    blanco = 'FFFFFF'

    borde = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws.merge_cells('A1:H1')
    ws['A1'] = 'ECUAMINERALES S.A.'
    ws['A1'].font = Font(bold=True, size=18, color=azul)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = 'Kardex de Inventario'
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:H3')
    ws['A3'] = f'Generado por: {request.user.username} | Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(size=10, italic=True)
    ws['A3'].alignment = Alignment(horizontal='center')

    encabezados = [
        'Fecha',
        'Usuario',
        'Insumo',
        'Tipo Movimiento',
        'Cantidad',
        'Stock Anterior',
        'Stock Actual',
        'Observación'
    ]

    fila_encabezado = 5

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True, color=blanco)
        celda.fill = PatternFill('solid', fgColor=azul)
        celda.alignment = Alignment(horizontal='center')
        celda.border = borde

    movimientos = MovimientoInventario.objects.all().order_by('-fecha_movimiento')

    fila = 6

    for mov in movimientos:

        datos = [
            mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M'),
            mov.usuario.username if mov.usuario else 'Sin usuario',
            mov.insumo.nombre_insumo if mov.insumo else 'Sin insumo',
            mov.tipo_movimiento,
            mov.cantidad,
            mov.stock_anterior,
            mov.stock_actual,
            mov.observacion
        ]

        for col, dato in enumerate(datos, start=1):
            celda = ws.cell(row=fila, column=col)
            celda.value = dato
            celda.border = borde
            celda.alignment = Alignment(vertical='center')

            if fila % 2 == 0:
                celda.fill = PatternFill('solid', fgColor=gris)

        tipo_celda = ws.cell(row=fila, column=4)

        if mov.tipo_movimiento == 'ENTRADA':
            tipo_celda.fill = PatternFill('solid', fgColor=verde)
            tipo_celda.font = Font(bold=True, color=blanco)
        else:
            tipo_celda.fill = PatternFill('solid', fgColor=rojo)
            tipo_celda.font = Font(bold=True, color=blanco)

        tipo_celda.alignment = Alignment(horizontal='center')

        fila += 1

    ws.auto_filter.ref = f'A5:H{fila - 1}'
    ws.freeze_panes = 'A6'

    anchos = {
        'A': 20,
        'B': 18,
        'C': 25,
        'D': 18,
        'E': 12,
        'F': 16,
        'G': 16,
        'H': 45,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="kardex_inventario.xlsx"'

    wb.save(response)

    return response