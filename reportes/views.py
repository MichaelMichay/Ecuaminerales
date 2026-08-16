from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from .models import Auditoria
import csv
from datetime import timedelta
from reportlab.lib import colors
from django.conf import settings
from datetime import datetime
import os
from despachos.models import OrdenTiro, Despacho

try:
    from polvorin.models import EntregaPolvorin
except:
    EntregaPolvorin = None

@login_required
def lista_auditoria(request):
    buscar = request.GET.get('buscar')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    auditorias = Auditoria.objects.all().order_by('-fecha')

    if buscar:
        auditorias = auditorias.filter(
            usuario__username__icontains=buscar
        ) | auditorias.filter(
            accion__icontains=buscar
        ) | auditorias.filter(
            descripcion__icontains=buscar
        )

    if fecha_inicio:
        auditorias = auditorias.filter(fecha__date__gte=fecha_inicio)

    if fecha_fin:
        auditorias = auditorias.filter(fecha__date__lte=fecha_fin)

    paginator = Paginator(auditorias, 10)
    page_number = request.GET.get('page')
    auditorias = paginator.get_page(page_number)

    return render(request, 'reportes/lista_reportes.html', {
        'auditorias': auditorias,
        'buscar': buscar,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'page_obj': auditorias
    })

@login_required
def reporte_auditoria_pdf(request):

    buscar = request.GET.get('buscar')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    auditorias = Auditoria.objects.all().order_by('-fecha')

    if buscar:
        auditorias = auditorias.filter(
            usuario__username__icontains=buscar
        ) | auditorias.filter(
            accion__icontains=buscar
        ) | auditorias.filter(
            descripcion__icontains=buscar
        )

    if fecha_inicio:
        auditorias = auditorias.filter(fecha__date__gte=fecha_inicio)

    if fecha_fin:
        auditorias = auditorias.filter(fecha__date__lte=fecha_fin)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_auditoria.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    def encabezado():

        logo_path = os.path.join(
            settings.BASE_DIR,
            'static',
            'img',
            'logo_ecuaminerales.jpeg'
        )

        if os.path.exists(logo_path):
            p.drawImage(
                logo_path,
                40,
                height - 90,
                width=75,
                height=55,
                preserveAspectRatio=True,
                mask='auto'
            )

        p.setFillColor(colors.HexColor('#243b80'))
        p.setFont('Helvetica-Bold', 22)

        p.drawCentredString(
            width / 2,
            height - 45,
            'ECUMINERALES S.A.'
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica-Bold', 15)

        p.drawCentredString(
            width / 2,
            height - 70,
            'Reporte de Auditoría del Sistema'
        )

        p.setFillColor(colors.HexColor('#555555'))
        p.setFont('Helvetica-Oblique', 9)

        p.drawCentredString(
            width / 2,
            height - 88,
            'Sistema de Control de Explosivos e Inventario'
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 9)

        p.drawString(
            40,
            height - 115,
            f'Generado por: {request.user.username}'
        )

        p.drawRightString(
            width - 40,
            height - 115,
            f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        )

        p.setStrokeColor(colors.HexColor('#243b80'))
        p.setLineWidth(1)
        p.line(40, height - 130, width - 40, height - 130)

    def cabecera_tabla(y):

        p.setFillColor(colors.HexColor('#243b80'))
        p.rect(35, y - 6, width - 70, 22, fill=True, stroke=False)

        p.setFillColor(colors.white)
        p.setFont('Helvetica-Bold', 7)

        p.drawString(40, y, 'FECHA')
        p.drawString(115, y, 'USUARIO')
        p.drawString(205, y, 'ACCIÓN')
        p.drawString(335, y, 'DESCRIPCIÓN')

        p.setFillColor(colors.black)

    def pie_pagina():

        p.setFont('Helvetica', 8)
        p.setFillColor(colors.grey)

        p.drawString(
            40,
            40,
            'ECUMINERALES S.A. - Sistema de control de explosivos e inventario'
        )

        p.drawRightString(
            width - 40,
            40,
            f'Página {p.getPageNumber()}'
        )

    def color_accion(accion):

        accion = accion.upper()

        if 'CREACIÓN' in accion:
            return '#0d6efd'

        if 'EDICIÓN' in accion:
            return '#ffc107'

        if 'ELIMINACIÓN' in accion:
            return '#dc3545'

        if 'APROBACIÓN' in accion:
            return '#198754'

        if 'RECHAZO' in accion:
            return '#dc3545'

        if 'ENTREGA' in accion:
            return '#6c757d'

        if 'CONTRASEÑA' in accion:
            return '#212529'

        if 'PERFIL' in accion:
            return '#0dcaf0'

        return '#6c757d'

    encabezado()

    y = height - 165
    cabecera_tabla(y)
    y -= 28

    contador = 0

    for auditoria in auditorias:

        if contador % 2 == 0:
            p.setFillColor(colors.HexColor('#f2f4f8'))
            p.rect(35, y - 5, width - 70, 22, fill=True, stroke=False)

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 7)

        p.drawString(
            40,
            y,
            auditoria.fecha.strftime('%d/%m/%Y %H:%M')
        )

        usuario = auditoria.usuario.username if auditoria.usuario else 'Sin usuario'

        p.drawString(
            115,
            y,
            usuario[:16]
        )

        accion = str(auditoria.accion)

        color = color_accion(accion)

        p.setFillColor(colors.HexColor(color))
        p.roundRect(205, y - 4, 115, 14, 4, fill=True, stroke=False)

        if color == '#ffc107' or color == '#0dcaf0':
            p.setFillColor(colors.black)
        else:
            p.setFillColor(colors.white)

        p.setFont('Helvetica-Bold', 6)
        p.drawCentredString(
            262,
            y,
            accion[:24]
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 7)

        descripcion = str(auditoria.descripcion)

        p.drawString(
            335,
            y,
            descripcion[:65]
        )

        y -= 24
        contador += 1

        if y < 80:

            pie_pagina()
            p.showPage()

            encabezado()

            y = height - 165
            cabecera_tabla(y)
            y -= 28

    pie_pagina()

    p.showPage()
    p.save()

    return response

@login_required
def trazabilidad_operacional(request):

    buscar = request.GET.get('buscar')

    ordenes = OrdenTiro.objects.all().order_by('-fecha_orden')

    if buscar:
        ordenes = ordenes.filter(codigo_orden__icontains=buscar)

    trazabilidad = []

    for orden in ordenes:

        despacho = Despacho.objects.filter(
            orden_tiro=orden
        ).first()

        entrega = None

        if despacho and EntregaPolvorin:
            entrega = EntregaPolvorin.objects.filter(
                despacho=despacho
            ).first()

        trazabilidad.append({
            'orden': orden,
            'despacho': despacho,
            'entrega': entrega
        })

    paginator = Paginator(trazabilidad, 10)
    page_number = request.GET.get('page')
    trazabilidad = paginator.get_page(page_number)

    return render(request, 'reportes/trazabilidad.html', {
        'trazabilidad': trazabilidad,
        'buscar': buscar,
        'page_obj': trazabilidad
    })

@login_required
def trazabilidad_pdf(request):

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="trazabilidad_operacional.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    ordenes = OrdenTiro.objects.all().order_by('-fecha_orden')

    def encabezado():

        logo_path = os.path.join(
            settings.BASE_DIR,
            'static',
            'img',
            'logo_ecuaminerales.jpeg'
        )

        if os.path.exists(logo_path):
            p.drawImage(
                logo_path,
                40,
                height - 90,
                width=75,
                height=55,
                preserveAspectRatio=True,
                mask='auto'
            )

        p.setFillColor(colors.HexColor('#243b80'))
        p.setFont('Helvetica-Bold', 22)

        p.drawCentredString(
            width / 2,
            height - 45,
            'ECUAMINERALES S.A.'
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica-Bold', 15)

        p.drawCentredString(
            width / 2,
            height - 70,
            'Trazabilidad Operacional'
        )

        p.setFillColor(colors.HexColor('#555555'))
        p.setFont('Helvetica-Oblique', 9)

        p.drawCentredString(
            width / 2,
            height - 88,
            'Seguimiento desde la solicitud hasta la entrega física'
        )

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 9)

        p.drawString(
            40,
            height - 115,
            f'Generado por: {request.user.username}'
        )

        p.drawRightString(
            width - 40,
            height - 115,
            f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        )

        p.setStrokeColor(colors.HexColor('#243b80'))
        p.setLineWidth(1)
        p.line(40, height - 130, width - 40, height - 130)

    def cabecera_tabla(y):

        p.setFillColor(colors.HexColor('#243b80'))
        p.rect(35, y - 6, width - 70, 22, fill=True, stroke=False)

        p.setFillColor(colors.white)
        p.setFont('Helvetica-Bold', 7)

        p.drawString(40, y, 'ORDEN')
        p.drawString(105, y, 'FECHA')
        p.drawString(190, y, 'PERFORISTA')
        p.drawString(285, y, 'BODEGUERO')
        p.drawString(380, y, 'POLVORINERO')
        p.drawString(490, y, 'ESTADO')

        p.setFillColor(colors.black)

    def pie_pagina():

        p.setFont('Helvetica', 8)
        p.setFillColor(colors.grey)

        p.drawString(
            40,
            40,
            'ECUMINERALES S.A. - Sistema de control de explosivos e inventario'
        )

        p.drawRightString(
            width - 40,
            40,
            f'Página {p.getPageNumber()}'
        )

    encabezado()

    y = height - 165
    cabecera_tabla(y)
    y -= 28

    contador = 0

    for orden in ordenes:

        despacho = Despacho.objects.filter(
            orden_tiro=orden
        ).first()

        entrega = None

        if despacho and EntregaPolvorin:
            entrega = EntregaPolvorin.objects.filter(
                despacho=despacho
            ).first()

        bodeguero = despacho.bodeguero.username if despacho else 'Pendiente'
        polvorinero = entrega.polvorinero.username if entrega else 'Sin entrega'

        if entrega and entrega.estado == 'ENTREGADO':
            estado = 'COMPLETADO'
            color_estado = '#198754'

        elif entrega and entrega.estado == 'RECHAZADO':
            estado = 'RECHAZADO'
            color_estado = '#dc3545'

        elif despacho:
            estado = 'DESPACHADO'
            color_estado = '#0dcaf0'

        else:
            estado = 'PENDIENTE'
            color_estado = '#ffc107'

        if contador % 2 == 0:
            p.setFillColor(colors.HexColor('#f2f4f8'))
            p.rect(35, y - 5, width - 70, 22, fill=True, stroke=False)

        p.setFillColor(colors.black)
        p.setFont('Helvetica', 7)

        p.drawString(40, y, str(orden.codigo_orden)[:12])
        p.drawString(105, y, orden.fecha_orden.strftime('%d/%m/%Y %H:%M'))
        p.drawString(190, y, str(orden.perforista.username)[:14])
        p.drawString(285, y, str(bodeguero)[:14])
        p.drawString(380, y, str(polvorinero)[:14])

        p.setFillColor(colors.HexColor(color_estado))
        p.roundRect(490, y - 4, 80, 14, 4, fill=True, stroke=False)

        if estado in ['PENDIENTE', 'DESPACHADO']:
            p.setFillColor(colors.black)
        else:
            p.setFillColor(colors.white)

        p.setFont('Helvetica-Bold', 6)
        p.drawCentredString(530, y, estado)

        y -= 24
        contador += 1

        if y < 80:

            pie_pagina()
            p.showPage()

            encabezado()

            y = height - 165
            cabecera_tabla(y)
            y -= 28

    pie_pagina()

    p.showPage()
    p.save()

    return response

@login_required
def trazabilidad_excel(request):

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from django.http import HttpResponse
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = 'Trazabilidad'

    azul = '243B80'
    verde = '198754'
    rojo = 'DC3545'
    amarillo = 'FFC107'
    celeste = '0DCAF0'
    gris = 'F2F4F8'
    blanco = 'FFFFFF'
    negro = '000000'

    borde = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = 'ECUAMINERALES S.A.'
    ws['A1'].font = Font(bold=True, size=18, color=azul)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = 'Trazabilidad Operacional'
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:I3')
    ws['A3'] = (
        f'Generado por: {request.user.username} | '
        f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )
    ws['A3'].font = Font(size=10, italic=True)
    ws['A3'].alignment = Alignment(horizontal='center')

    encabezados = [
        'Orden',
        'Fecha Solicitud',
        'Perforista',
        'Cantidad Tiros',
        'Bodeguero',
        'Fecha Despacho',
        'Polvorinero',
        'Fecha Entrega',
        'Estado Final'
    ]

    fila_encabezado = 5

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True, color=blanco)
        celda.fill = PatternFill('solid', fgColor=azul)
        celda.alignment = Alignment(horizontal='center')
        celda.border = borde

    ordenes = OrdenTiro.objects.all().order_by('-fecha_orden')

    fila = 6

    for orden in ordenes:

        despacho = Despacho.objects.filter(
            orden_tiro=orden
        ).first()

        entrega = None

        if despacho and EntregaPolvorin:
            entrega = EntregaPolvorin.objects.filter(
                despacho=despacho
            ).first()

        estado_final = 'PENDIENTE'

        if despacho:
            estado_final = 'DESPACHADO'

        if entrega and entrega.estado == 'ENTREGADO':
            estado_final = 'COMPLETADO'

        if entrega and entrega.estado == 'RECHAZADO':
            estado_final = 'RECHAZADO EN POLVORÍN'

        datos = [
            orden.codigo_orden,
            orden.fecha_orden.strftime('%d/%m/%Y %H:%M'),
            orden.perforista.username,
            orden.cantidad_tiros,
            despacho.bodeguero.username if despacho else 'Pendiente',
            despacho.fecha_despacho.strftime('%d/%m/%Y %H:%M') if despacho else '',
            entrega.polvorinero.username if entrega else 'Sin entrega',
            entrega.fecha_entrega.strftime('%d/%m/%Y %H:%M') if entrega else '',
            estado_final
        ]

        for col, dato in enumerate(datos, start=1):
            celda = ws.cell(row=fila, column=col)
            celda.value = dato
            celda.border = borde
            celda.alignment = Alignment(vertical='center')

            if fila % 2 == 0:
                celda.fill = PatternFill('solid', fgColor=gris)

        estado_celda = ws.cell(row=fila, column=9)

        if estado_final == 'COMPLETADO':
            estado_celda.fill = PatternFill('solid', fgColor=verde)
            estado_celda.font = Font(bold=True, color=blanco)

        elif estado_final == 'RECHAZADO EN POLVORÍN':
            estado_celda.fill = PatternFill('solid', fgColor=rojo)
            estado_celda.font = Font(bold=True, color=blanco)

        elif estado_final == 'DESPACHADO':
            estado_celda.fill = PatternFill('solid', fgColor=celeste)
            estado_celda.font = Font(bold=True, color=negro)

        else:
            estado_celda.fill = PatternFill('solid', fgColor=amarillo)
            estado_celda.font = Font(bold=True, color=negro)

        estado_celda.alignment = Alignment(horizontal='center')

        fila += 1

    ws.auto_filter.ref = f'A5:I{fila - 1}'
    ws.freeze_panes = 'A6'

    anchos = {
        'A': 14,
        'B': 20,
        'C': 18,
        'D': 16,
        'E': 18,
        'F': 20,
        'G': 18,
        'H': 20,
        'I': 22,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="trazabilidad_operacional.xlsx"'

    wb.save(response)

    return response